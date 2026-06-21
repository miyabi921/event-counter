from flask import Flask, jsonify, request, send_file
import json, os, io, threading, socket
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
lock = threading.Lock()

# ── Storage: Redis（クラウド）or ローカルJSONファイル ──────────────────────────
REDIS_URL = os.environ.get("REDIS_URL", "")
_redis = None
if REDIS_URL:
    try:
        import redis as redis_lib
        _redis = redis_lib.from_url(REDIS_URL, decode_responses=True)
        _redis.ping()
        print("Redis connected.")
    except Exception as e:
        print(f"Redis connection failed: {e}. Falling back to local file.")
        _redis = None

CATEGORIES = [
    "AIサイネージ",
    "リアルタイムキャスト",
    "AIカメラ分析",
    "コンテンツ自動生成",
    "災害情報連携・配信",
    "ポッドル（チラシ音声自動生成）",
    "タブレットミニマムプラン",
]

CAT_COLORS = {
    "AIサイネージ":                  {"bg": "#D6EAF8", "accent": "#1B4F72", "icon": "📺"},
    "リアルタイムキャスト":          {"bg": "#D5F5E3", "accent": "#1E8449", "icon": "📡"},
    "AIカメラ分析":                  {"bg": "#E8DAEF", "accent": "#6C3483", "icon": "📷"},
    "コンテンツ自動生成":            {"bg": "#FDEBD0", "accent": "#B7410E", "icon": "✨"},
    "災害情報連携・配信":            {"bg": "#FADBD8", "accent": "#C0392B", "icon": "🚨"},
    "ポッドル（チラシ音声自動生成）": {"bg": "#D1F2EB", "accent": "#117A65", "icon": "🎙️"},
    "タブレットミニマムプラン":      {"bg": "#FEF9E7", "accent": "#7D6608", "icon": "📱"},
}

CAT_FILL_HEX = {
    "AIサイネージ":                  "D6EAF8",
    "リアルタイムキャスト":          "D5F5E3",
    "AIカメラ分析":                  "E8DAEF",
    "コンテンツ自動生成":            "FDEBD0",
    "災害情報連携・配信":            "FADBD8",
    "ポッドル（チラシ音声自動生成）": "D1F2EB",
    "タブレットミニマムプラン":      "FEF9E7",
}

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "counts.json")


def load_data():
    # Redisから読む
    if _redis:
        try:
            raw = _redis.get("event_data")
            if raw:
                return json.loads(raw)
        except Exception as e:
            print(f"Redis read error: {e}")
    # ローカルファイルから読む
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"totals": {cat: 0 for cat in CATEGORIES}, "log": []}


def save_data(d):
    serialized = json.dumps(d, ensure_ascii=False)
    if _redis:
        try:
            _redis.set("event_data", serialized)
            return
        except Exception as e:
            print(f"Redis write error: {e}")
    # ローカルファイルに保存
    try:
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            f.write(serialized)
    except Exception as e:
        print(f"File write error: {e}")


data = load_data()

# ── Counter page HTML ─────────────────────────────────────────────────────────

COUNTER_HTML = r"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, user-scalable=no">
<title>展示会カウンター</title>
<style>
*{box-sizing:border-box;margin:0;padding:0;-webkit-tap-highlight-color:transparent}
body{font-family:'Hiragino Sans','Yu Gothic',sans-serif;background:#1a2332;min-height:100vh;display:flex;flex-direction:column}
.header{background:#2E4057;color:#fff;padding:14px 18px;display:flex;justify-content:space-between;align-items:center}
.header h1{font-size:1.1em;font-weight:700}
.sync{font-size:.75em;display:flex;align-items:center;gap:6px;color:#aac}
.dot{width:8px;height:8px;border-radius:50%;background:#2ecc71;display:inline-block}
.dot.old{background:#e74c3c}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:12px;padding:14px;flex:1}
.card{border-radius:16px;overflow:hidden;user-select:none;display:flex;flex-direction:column;position:relative}
.card-info{display:flex;flex-direction:column;align-items:center;padding:16px 12px 12px;gap:4px}
.icon{font-size:2em;line-height:1}
.name{font-size:.82em;font-weight:700;text-align:center;line-height:1.35}
.count-wrap{display:flex;flex-direction:column;align-items:center;margin:6px 0 2px}
.count{font-size:3em;font-weight:800;line-height:1;transition:transform .12s}
.count.bump{transform:scale(1.28)}
.unit{font-size:.72em;font-weight:600;opacity:.7;margin-top:2px}
.card-btns{display:grid;grid-template-columns:1fr 1fr;border-top:2px solid rgba(0,0,0,.12)}
.btn-minus,.btn-plus{border:none;cursor:pointer;padding:14px 0;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:3px;touch-action:manipulation;transition:filter .1s;-webkit-tap-highlight-color:transparent}
.btn-minus:active,.btn-plus:active{filter:brightness(.82)}
.btn-minus{border-right:1px solid rgba(0,0,0,.1)}
.btn-sym{font-size:1.8em;font-weight:800;line-height:1}
.btn-lbl{font-size:.68em;font-weight:600;opacity:.8}
.ripple{position:absolute;top:35%;left:50%;transform:translate(-50%,-50%);font-size:1.6em;font-weight:800;pointer-events:none;animation:floatup .6s ease-out forwards}
@keyframes floatup{0%{transform:translate(-50%,-50%) scale(1);opacity:1}100%{transform:translate(-50%,-130%) scale(1.4);opacity:0}}
.footer{background:#2E4057;color:#aac;font-size:.72em;text-align:center;padding:8px;display:flex;justify-content:center;gap:16px}
.footer a{color:#88aacc;text-decoration:none}
</style>
</head>
<body>
<div class="header">
  <h1>🎯 展示会カウンター</h1>
  <div class="sync"><span class="dot" id="dot"></span><span id="syncLabel">同期中...</span></div>
</div>
<div class="grid" id="grid"></div>
<div class="footer">
  <span id="total">合計 0 件</span>
  <span>｜</span>
  <a href="/admin">📊 集計・管理画面</a>
</div>

<script>
const CATS=__CATS__;
const COLORS=__COLORS__;
let counts={};
let lastSync=0;

function init(){
  const g=document.getElementById('grid');
  CATS.forEach(cat=>{
    const c=COLORS[cat];
    const card=document.createElement('div');
    card.className='card';
    card.id='card-'+cat;
    card.style.background=c.bg;

    const minusBg=`rgba(192,57,43,0.15)`;
    const plusBg=`rgba(0,0,0,0.08)`;

    card.innerHTML=`
      <div class="card-info">
        <span class="icon">${c.icon}</span>
        <span class="name" style="color:${c.accent}">${cat}</span>
        <div class="count-wrap">
          <span class="count" id="cnt-${cat}" style="color:${c.accent}">0</span>
          <span class="unit" style="color:${c.accent}">件</span>
        </div>
      </div>
      <div class="card-btns">
        <button class="btn-minus" style="background:${minusBg};color:#C0392B"
          ontouchstart="tap(event,'${cat}',-1)" onclick="tap(event,'${cat}',-1)">
          <span class="btn-sym">−</span>
          <span class="btn-lbl">取消</span>
        </button>
        <button class="btn-plus" style="background:${plusBg};color:${c.accent}"
          ontouchstart="tap(event,'${cat}',1)" onclick="tap(event,'${cat}',1)">
          <span class="btn-sym">＋</span>
          <span class="btn-lbl">カウント</span>
        </button>
      </div>`;
    g.appendChild(card);
  });
  refresh();
  setInterval(refresh, 2000);
}

function tap(e, cat, delta){
  e.stopPropagation();
  // Optimistic UI update
  counts[cat]=(counts[cat]||0)+delta;
  if(counts[cat]<0)counts[cat]=0;
  updateDisplay();
  // Animate
  if(delta>0){
    const el=document.getElementById('cnt-'+cat);
    el.classList.add('bump');
    setTimeout(()=>el.classList.remove('bump'),160);
    const info=document.querySelector('#card-'+cat+' .card-info');
    const rip=document.createElement('span');
    rip.className='ripple';
    rip.style.color=COLORS[cat].accent;
    rip.textContent='+1';
    info.appendChild(rip);
    setTimeout(()=>rip.remove(),620);
  }
  // Send to server
  fetch('/count',{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({category:cat,delta:delta})
  }).then(r=>r.json()).then(d=>{
    if(d.totals)counts=d.totals;
    updateDisplay();
  }).catch(()=>{});
}

function refresh(){
  fetch('/results')
    .then(r=>r.json())
    .then(d=>{
      counts=d.totals||{};
      updateDisplay();
      lastSync=Date.now();
      document.getElementById('dot').className='dot';
      document.getElementById('syncLabel').textContent='同期中 ✓';
    })
    .catch(()=>{
      const age=Date.now()-lastSync;
      document.getElementById('dot').className='dot old';
      document.getElementById('syncLabel').textContent='オフライン';
    });
}

function updateDisplay(){
  let total=0;
  CATS.forEach(cat=>{
    const n=counts[cat]||0;
    document.getElementById('cnt-'+cat).textContent=n;
    total+=n;
  });
  document.getElementById('total').textContent='合計 '+total+' 件';
}

init();
</script>
</body>
</html>"""

# ── Admin page HTML ───────────────────────────────────────────────────────────

ADMIN_HTML = r"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>集計・管理画面</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Hiragino Sans','Yu Gothic',sans-serif;background:#f0f4f8;min-height:100vh}
.header{background:#2E4057;color:#fff;padding:16px 20px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px}
.header h1{font-size:1.2em;font-weight:700}
.btns{display:flex;gap:10px;flex-wrap:wrap}
.btn{padding:10px 18px;border-radius:8px;border:none;cursor:pointer;font-family:inherit;font-size:.9em;font-weight:700;transition:opacity .15s}
.btn:hover{opacity:.82}
.export{background:#27AE60;color:#fff}
.reset{background:#E74C3C;color:#fff}
.back{background:#555;color:#fff}
.summary{background:#fff;margin:16px;padding:16px 20px;border-radius:12px;display:flex;gap:28px;flex-wrap:wrap;box-shadow:0 2px 8px rgba(0,0,0,.07)}
.si .num{font-size:2em;font-weight:800;color:#2E4057}
.si .lbl{font-size:.76em;color:#888;margin-top:2px}
.cards{padding:0 16px 20px;display:flex;flex-direction:column;gap:10px;max-width:820px;margin:0 auto}
.rc{background:#fff;border-radius:12px;padding:16px 18px;box-shadow:0 2px 8px rgba(0,0,0,.07)}
.rh{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px}
.rname{font-weight:700;font-size:1em}
.rcount{font-size:1.5em;font-weight:800}
.barbg{background:#eee;border-radius:6px;height:12px;overflow:hidden}
.barfill{height:100%;border-radius:6px;transition:width .5s ease}
.note{text-align:center;color:#bbb;font-size:.75em;padding:8px 0 16px}
</style>
</head>
<body>
<div class="header">
  <h1>📊 集計・管理画面</h1>
  <div class="btns">
    <a class="btn back" href="/">← カウンターへ戻る</a>
    <button class="btn export" onclick="location.href='/export'">📥 Excelに出力</button>
    <button class="btn reset" onclick="doReset()">🔄 リセット</button>
  </div>
</div>
<div class="summary">
  <div class="si"><div class="num" id="nTotal">-</div><div class="lbl">カウント合計</div></div>
  <div class="si"><div class="num" id="nTop">-</div><div class="lbl">1位カテゴリ</div></div>
  <div class="si"><div class="num" id="nUpd">-</div><div class="lbl">最終更新</div></div>
</div>
<div class="cards" id="cards"></div>
<div class="note">3秒ごとに自動更新</div>

<script>
const CATS=__CATS__;
const COLORS=__COLORS__;
function render(d){
  const tot=d.totals||{};
  const total=Object.values(tot).reduce((a,b)=>a+b,0);
  const max=Math.max(...Object.values(tot),1);
  const topCat=[...CATS].sort((a,b)=>(tot[b]||0)-(tot[a]||0))[0];
  document.getElementById('nTotal').textContent=total+'件';
  document.getElementById('nTop').textContent=(tot[topCat]||0)>0?COLORS[topCat].icon+' '+(topCat.length>6?topCat.slice(0,6)+'…':topCat):'-';
  document.getElementById('nUpd').textContent=new Date().toLocaleTimeString('ja-JP');
  const sorted=[...CATS].sort((a,b)=>(tot[b]||0)-(tot[a]||0));
  const pct=cat=>total>0?Math.round((tot[cat]||0)/total*100):0;
  document.getElementById('cards').innerHTML=sorted.map(cat=>{
    const n=tot[cat]||0,c=COLORS[cat];
    const bar=Math.round(n/max*100);
    return `<div class="rc">
      <div class="rh"><span class="rname">${c.icon} ${cat}</span><span class="rcount" style="color:${c.accent}">${n} 件</span></div>
      <div class="barbg"><div class="barfill" style="width:${bar}%;background:${c.accent}"></div></div>
      <div style="font-size:.76em;color:#999;margin-top:4px">全体の ${pct(cat)}%</div>
    </div>`;
  }).join('');
}
function doReset(){
  if(!confirm('全カウントをリセットしますか？\nExcelへの出力が済んでいることを確認してください。'))return;
  fetch('/reset',{method:'POST'}).then(r=>r.json()).then(d=>{if(d.ok)refresh();});
}
function refresh(){fetch('/results').then(r=>r.json()).then(render);}
refresh();
setInterval(refresh,3000);
</script>
</body>
</html>"""


# ── Helpers ───────────────────────────────────────────────────────────────────

def inject(template):
    cats = json.dumps(CATEGORIES, ensure_ascii=False)
    colors = json.dumps(CAT_COLORS, ensure_ascii=False)
    return template.replace("__CATS__", cats).replace("__COLORS__", colors)


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return inject(COUNTER_HTML)


@app.route("/admin")
def admin():
    return inject(ADMIN_HTML)


@app.route("/count", methods=["POST"])
def count():
    global data
    body = request.get_json(silent=True) or {}
    cat = body.get("category", "")
    delta = int(body.get("delta", 1))
    if cat not in CATEGORIES:
        return jsonify({"error": "unknown category"}), 400
    with lock:
        current = data["totals"].get(cat, 0) + delta
        data["totals"][cat] = max(0, current)
        if delta > 0:
            data["log"].append({
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "category": cat,
                "delta": delta,
            })
        save_data(data)
    return jsonify({"ok": True, "totals": data["totals"]})


@app.route("/results")
def results():
    return jsonify({"totals": data["totals"], "log_count": len(data.get("log", []))})


@app.route("/reset", methods=["POST"])
def reset():
    global data
    with lock:
        data = {"totals": {cat: 0 for cat in CATEGORIES}, "log": []}
        save_data(data)
    return jsonify({"ok": True})


@app.route("/export")
def export():
    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="CCCCCC")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Sheet 1: 集計サマリー ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "集計サマリー"

    ws1.merge_cells("A1:D1")
    ws1["A1"] = "展示会 興味カウント 集計結果"
    ws1["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", start_color="2E4057")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 38

    total_count = sum(data["totals"].values())
    ws1.merge_cells("A2:D2")
    ws1["A2"] = f"出力日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}　／　カウント合計: {total_count} 件"
    ws1["A2"].font = Font(name="Arial", size=10, color="888888")
    ws1.row_dimensions[2].height = 20

    for col, h in enumerate(["カテゴリ", "カウント数", "全体比 (%)", "順位"], 1):
        c = ws1.cell(row=4, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", start_color="34495E")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
    ws1.row_dimensions[4].height = 28

    sorted_cats = sorted(CATEGORIES, key=lambda x: data["totals"].get(x, 0), reverse=True)
    for rank, cat in enumerate(sorted_cats, 1):
        n = data["totals"].get(cat, 0)
        pct = round(n / total_count * 100, 1) if total_count > 0 else 0
        row = rank + 4
        fill = PatternFill("solid", start_color=CAT_FILL_HEX.get(cat, "FFFFFF"))
        for col, val in enumerate([cat, n, pct, rank], 1):
            c = ws1.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = fill
            c.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
            c.border = bdr
        ws1.row_dimensions[row].height = 24

    ws1.column_dimensions["A"].width = 34
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 8

    # ── Sheet 2: タイムライン ────────────────────────────────────────────────
    ws2 = wb.create_sheet("タイムライン")
    for col, h in enumerate(["NO", "日時", "カテゴリ"], 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", start_color="2E4057")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
    ws2.row_dimensions[1].height = 28

    for i, entry in enumerate(data.get("log", []), 1):
        row = i + 1
        cat = entry.get("category", "")
        fill = PatternFill("solid", start_color=CAT_FILL_HEX.get(cat, "FFFFFF"))
        for col, val in enumerate([i, entry.get("timestamp", ""), cat], 1):
            c = ws2.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = fill
            c.alignment = Alignment(horizontal="center" if col != 3 else "left", vertical="center")
            c.border = bdr
        ws2.row_dimensions[row].height = 20

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 34
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"集計結果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Startup ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    public_url = None
    ngrok_token = os.environ.get("NGROK_TOKEN", "").strip()

    if ngrok_token:
        try:
            from pyngrok import ngrok, conf
            conf.get_default().auth_token = ngrok_token
            tunnel = ngrok.connect(5000, bind_tls=True)
            public_url = tunnel.public_url
        except ImportError:
            print("[警告] pyngrok がインストールされていません。pip install pyngrok を実行してください。")
        except Exception as e:
            print(f"[警告] ngrok 起動失敗: {e}")

    print("=" * 60)
    print("  展示会カウンターアプリ 起動中")
    print("=" * 60)
    if public_url:
        print(f"  【外部公開URL（LTE可）】")
        print(f"  カウンター  →  {public_url}/")
        print(f"  管理画面    →  {public_url}/admin")
    else:
        try:
            local_ip = socket.gethostbyname(socket.gethostname())
        except Exception:
            local_ip = "127.0.0.1"
        print(f"  カウンター  →  http://{local_ip}:5000/")
        print(f"  管理画面    →  http://{local_ip}:5000/admin")
        print()
        print("  ※ LTE対応は 外部公開_LTE対応_起動.bat を使用")
    print("=" * 60)
    print("  Ctrl+C で終了")
    print("=" * 60)

    app.run(host="0.0.0.0", port=5000, debug=False)
